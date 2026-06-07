import openpyxl
import re
import os
import warnings
from pyproj import Proj

# Matiin warning remeh dari openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def process_and_route(input_file, output_excel, output_folder):
    utm_proj = Proj(proj='utm', zone=47, south=True, ellps='WGS84')
    
    print(f"Reading {input_file}...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    pole_data = {}
    pole_list = []
    
    # --- SETUP HEADER OUTPUT ---
    ws.cell(row=1, column=22).value = "Input_Kabel"         
    ws.cell(row=1, column=23).value = "Input_tiang"         
    ws.cell(row=1, column=24).value = "Input_Umbang"        
    ws.cell(row=1, column=25).value = "Input_No Tiang"      
    ws.cell(row=1, column=26).value = "Input_No Tiang Lama" 
    ws.cell(row=1, column=27).value = "Input_details_Summary" 
    
    cable_headers = {}
    for col_idx in range(12, 22):
        val = ws.cell(row=1, column=col_idx).value
        cable_headers[col_idx] = str(val).strip() if val else f"Cable_{col_idx}"
        
    def get_cable_layer(cable_name):
        name_up = str(cable_name).upper()
        if "185" in name_up: return "CABLE 185"
        if "95" in name_up: return "CABLE 95"
        if "3 X 16" in name_up or "3X16" in name_up: return "CABLE 316"
        if "1X 16" in name_up or "1X16" in name_up: return "CABLE SERVICE"
        if "19064" in name_up or "9064" in name_up: return "PVC 9064"
        if "7083" in name_up or "7044" in name_up: return "PVC 7083"
        if "7173" in name_up or "3132" in name_up: return "BARE 7173"
        if "7122" in name_up: return "BARE 7122"
        return "CABLE SERVICE" 
    
    # --- PASS 1: Read and Clean Data ---
    for row in range(2, ws.max_row + 1):  
        coord_str = ws.cell(row=row, column=1).value
        main_pole_name = ws.cell(row=row, column=2).value
        
        if not coord_str or not main_pole_name:
            continue
            
        try:
            lat_str, lon_str = str(coord_str).split(',')
            lat = float(lat_str.strip())
            lon = float(lon_str.strip())
            x, y = utm_proj(lon, lat)
            
            pole_name_clean = str(main_pole_name).strip()
            
            old_pole_val = ws.cell(row=row, column=3).value
            old_pole_name = str(old_pole_val).strip() if old_pole_val else ""
            
            umbang_raw = ws.cell(row=row, column=4).value
            umbang_val = str(umbang_raw).strip() if umbang_raw is not None else ""
            umbang_block = f"U{umbang_val}" if umbang_val.isdigit() and 1 <= int(umbang_val) <= 9 else None
            
            val_160 = ws.cell(row=row, column=5).value
            qty_bk = 0
            if val_160 is not None and str(val_160).strip() != "":
                try: qty_bk = int(float(str(val_160).strip()))
                except: pass
                
            val_400 = ws.cell(row=row, column=6).value
            qty_bm = 0
            if val_400 is not None and str(val_400).strip() != "":
                try: qty_bm = int(float(str(val_400).strip()))
                except: pass
                
            val_srv = ws.cell(row=row, column=7).value
            srv_text = ""
            if val_srv is not None and str(val_srv).strip() != "":
                srv_text = str(val_srv).strip()
                if srv_text.endswith('.0'): srv_text = srv_text[:-2] 
                
            details_summary = f"BK:{qty_bk}, BM:{qty_bm}, Srv:{srv_text}" if (qty_bk>0 or qty_bm>0 or srv_text) else ""
            
            val_h = str(ws.cell(row=row, column=8).value).strip()
            val_i = str(ws.cell(row=row, column=9).value).strip()
            val_j = str(ws.cell(row=row, column=10).value).strip()
            val_k = str(ws.cell(row=row, column=11).value).strip()
            
            active_flags = []
            if val_h in ["1", "1.0"]: active_flags.append("TIANG_MERAH")
            if val_i in ["1", "1.0"]: active_flags.append("TIANG_BIRU")
            if val_j in ["1", "1.0"]: active_flags.append("TIANG_MAG")
            if val_k in ["1", "1.0"]: active_flags.append("TIANG_SP")
            block_name = active_flags[0] if len(active_flags) == 1 else "TIANG_BIRU"
            
            dominant_found = False
            cable_name_for_layer = "Cable 1X 16 Nmp" 
            
            for col_idx in range(12, 22):
                cable_val = ws.cell(row=row, column=col_idx).value
                if not dominant_found and cable_val is not None and str(cable_val).strip() != "":
                    dominant_found = True
                    cable_name_for_layer = cable_headers[col_idx]
                elif dominant_found:
                    ws.cell(row=row, column=col_idx).value = None
            
            if not dominant_found:
                ws.cell(row=row, column=15).value = 1 
            
            final_layer_kabel = get_cable_layer(cable_name_for_layer)
            
            pole_data[pole_name_clean] = {
                'x': x, 'y': y, 'row': row, 
                'block': block_name, 'umbang': umbang_block,
                'old_pole': old_pole_name,
                'cable_layer': final_layer_kabel,
                'qty_bk': qty_bk,
                'qty_bm': qty_bm,
                'srv': srv_text,
                'summary': details_summary
            }
            pole_list.append(pole_name_clean)
            
        except Exception:
            continue

    # --- PASS 2: Execute Routing Logic ---
    pole_set = set(pole_list)
    pole_relations = {}
    
    def get_joint_pole(missing_target):
        m = re.search(r'^(.*?\s+)([A-Z]+)\s+(\d+)$', missing_target)
        if m:
            prefix = m.group(1)
            feeder = m.group(2)
            num = m.group(3)
            for t in pole_set:
                m_t = re.search(r'^(.*?\s+)([A-Z]+)\s+' + num + r'$', t)
                if m_t and prefix == m_t.group(1) and feeder in m_t.group(2):
                    return t
        return None

    for pole in pole_list:
        parent = None
        match_suffix_child = re.search(r'^(.*)\s+/(\d+)([A-Za-z]+)$', pole)
        match_child = re.search(r'^(.*)\s+/(\d+)$', pole)
        match_main = re.search(r'^(.*?\s+)([A-Z]+)\s+(\d+)$', pole)
        
        if match_suffix_child:
            base = match_suffix_child.group(1).strip()
            num = int(match_suffix_child.group(2))
            suffix = match_suffix_child.group(3)
            if num != 1:
                target_parent = f"{base} /{num - 1}{suffix}"
                parent = target_parent if target_parent in pole_set else (get_joint_pole(target_parent) or target_parent)
            else:
                target_parent = base
                parent = target_parent if target_parent in pole_set else (get_joint_pole(target_parent) or target_parent)
        elif match_child:
            base = match_child.group(1).strip()
            num = int(match_child.group(2))
            if num != 1:
                target_parent = f"{base} /{num - 1}"
                parent = target_parent if target_parent in pole_set else target_parent
            else:
                target_parent = base
                parent = target_parent if target_parent in pole_set else (get_joint_pole(target_parent) or target_parent)
        elif match_main:
            prefix = match_main.group(1)
            feeder = match_main.group(2)
            num = int(match_main.group(3))
            if num != 1:
                target_parent = f"{prefix}{feeder} {num - 1}"
                parent = target_parent if target_parent in pole_set else (get_joint_pole(target_parent) or target_parent)
                
        pole_relations[pole] = parent

    # --- PASS 3: Generate Split AutoCAD Commands with Auto-Grouping ---
    os.makedirs(output_folder, exist_ok=True)
    
    scripts_by_area = {}
    current_area = None
    
    for pole in pole_list: 
        data = pole_data[pole]
        
        if pole.upper().startswith("PE "):
            current_area = pole
        elif current_area is None:
            current_area = pole
            
        clean_area_name = re.sub(r'[\\/*?:"<>|]', "", current_area).strip()
        
        if clean_area_name not in scripts_by_area:
            scripts_by_area[clean_area_name] = [
                '(setvar "OSMODE" 0)',
                "(setq firstEnt (entlast))" 
            ]
            
        cmds = scripts_by_area[clean_area_name]
        
        row = data['row']
        x = data['x']
        y = data['y']
        block_name = data['block']
        umbang_block = data['umbang']
        old_pole_text = data['old_pole']
        cable_layer = data['cable_layer']
        
        qty_bk = data['qty_bk']
        qty_bm = data['qty_bm']
        srv_val = data['srv']
        
        # text_x = x + 2.0
        # y_pole_main = y + 1.5 
        # y_pole_lama = y + 0.0 
        
        # srv_x = x - 1.5
        # srv_y = y - 1.5
        
        # bb_start_y = y - 1.5      
        # offset_y_bb = 1.2         
        # offset_x_bb = 2.0         
        
        # x_bk = x + 1.5            
        # x_bm = x_bk + offset_x_bb if qty_bk > 0 else x_bk  
        
        # Teks utama (Agak tinggi)
        text_x = x + 3.0
        y_pole_main = y + 4.4 
        
        # Tiang lama / NA (Persis di bawah teks utama)
        y_pole_lama = y + 1.5 
        
        # Kordinat Service (Kiri Bawah)
        srv_x = x - 2.8
        srv_y = y - 3.2
        
        # SETUP KOORDINAT BLACKBOX (Kanan Bawah & Deret)
        bb_start_y = y - 3.2      
        offset_y_bb = 3.2         
        offset_x_bb = 5.2         
        
        x_bk = x + 3.5            
        x_bm = x_bk + offset_x_bb if qty_bk > 0 else x_bk


        # # 1. Line Routing Kabel
        # parent = pole_relations.get(pole)
        # if parent and parent in pole_data:
        #     x_parent = pole_data[parent]['x']
        #     y_parent = pole_data[parent]['y']
        #     cmd_line = f"LINE {x_parent:.3f},{y_parent:.3f} {x:.3f},{y:.3f} "
        #     ws.cell(row=row, column=22).value = cmd_line
        #     cmds.append(f'(command "-layer" "m" "{cable_layer}" "")')
        #     cmds.append(cmd_line)
            
        # # 2. Pole Block Utama
        # cmd_insert = f'(command "-insert" "{block_name}" "{x:.3f},{y:.3f}" 1 1 0)'
        # ws.cell(row=row, column=23).value = cmd_insert
        # cmds.append('(command "-layer" "m" "0" "")')
        # cmds.append(cmd_insert)

        # # 3. Umbang
        # if umbang_block:
        #     cmd_umbang = f'(command "-insert" "{umbang_block}" "{x:.3f},{y:.3f}" 1 1 0)'
        #     ws.cell(row=row, column=24).value = cmd_umbang
        #     cmds.append('(command "-layer" "m" "UMBANG" "")')
        #     cmds.append(cmd_umbang)
        
        # # 4. Teks Main Pole Name
        # if pole.upper().startswith("PE "):
        #     display_label_b = pole
        # else:
        #     display_name_match = re.search(r'([A-Z]+\s+\d+.*)$', pole)
        #     display_label_b = display_name_match.group(1) if display_name_match else pole 
            
        # if display_label_b: 
        #     cmd_text_1 = f"-TEXT {text_x:.3f},{y_pole_main:.3f} 0 {display_label_b}"
        #     ws.cell(row=row, column=25).value = cmd_text_1
        #     cmds.append('(command "-layer" "m" "POLE NUMBER" "")')
        #     cmds.append(cmd_text_1)
        
        # # 5. Teks Old Pole / N/A
        # if old_pole_text: 
        #     cmd_text_2 = f"-TEXT {text_x:.3f},{y_pole_lama:.3f} 0 {old_pole_text}"
        #     ws.cell(row=row, column=26).value = cmd_text_2
        #     cmds.append('(command "-layer" "m" "TIANG LAMA" "")')
        #     cmds.append(cmd_text_2)
            
        # ws.cell(row=row, column=27).value = data['summary']
        
        # # 6. INSERT BLOCK SERVICE (JUMLAH SERVIS)
        # if srv_val:
        #     block_service = f"S{srv_val}"
        #     cmd_srv = f'(command "-insert" "{block_service}" "{srv_x:.3f},{srv_y:.3f}" 1 1 0)'
        #     cmds.append('(command "-layer" "m" "JUMLAH SERVIS" "")')
        #     cmds.append(cmd_srv)
            
        # # 7. INSERT BLOCK BLACKBOX BK 160A (BBOX 160A)
        # for i in range(qty_bk):
        #     y_insert = bb_start_y - (i * offset_y_bb)
        #     cmd_bk = f'(command "-insert" "BK" "{x_bk:.3f},{y_insert:.3f}" 1 1 0)'
        #     cmds.append('(command "-layer" "m" "BBOX 160A" "")')
        #     cmds.append(cmd_bk)
            
        # # 8. INSERT BLOCK BLACKBOX BM 400A (BBOX 400A)
        # for i in range(qty_bm):
        #     y_insert = bb_start_y - (i * offset_y_bb)
        #     cmd_bm = f'(command "-insert" "BM" "{x_bm:.3f},{y_insert:.3f}" 1 1 0)'
        #     cmds.append('(command "-layer" "m" "BBOX 400A" "")')
        #     cmds.append(cmd_bm)

        # 1. Line Routing Kabel
        parent = pole_relations.get(pole)
        if parent and parent in pole_data:
            x_parent = pole_data[parent]['x']
            y_parent = pole_data[parent]['y']
            cmd_line = f"LINE {x_parent:.3f},{y_parent:.3f} {x:.3f},{y:.3f} "
            ws.cell(row=row, column=22).value = cmd_line
            cmds.append(f'(command "-layer" "m" "{cable_layer}" "")')
            cmds.append(cmd_line)

        # 6. INSERT BLOCK SERVICE (JUMLAH SERVIS) -> SEKARANG DI SINI (SEBELUM TIANG)
        if srv_val:
            block_service = f"S{srv_val}"
            cmd_srv = f'(command "-insert" "{block_service}" "{srv_x:.3f},{srv_y:.3f}" 1 1 0)'
            cmds.append('(command "-layer" "m" "JUMLAH SERVIS" "")')
            cmds.append(cmd_srv)
            
        # 2. Pole Block Utama
        cmd_insert = f'(command "-insert" "{block_name}" "{x:.3f},{y:.3f}" 1 1 0)'
        ws.cell(row=row, column=23).value = cmd_insert
        cmds.append('(command "-layer" "m" "0" "")')
        cmds.append(cmd_insert)

        # 3. Umbang
        if umbang_block:
            cmd_umbang = f'(command "-insert" "{umbang_block}" "{x:.3f},{y:.3f}" 1 1 0)'
            ws.cell(row=row, column=24).value = cmd_umbang
            cmds.append('(command "-layer" "m" "UMBANG" "")')
            cmds.append(cmd_umbang)
        
        # 4. Teks Main Pole Name
        if pole.upper().startswith("PE "):
            display_label_b = pole
        else:
            display_name_match = re.search(r'([A-Z]+\s+\d+.*)$', pole)
            display_label_b = display_name_match.group(1) if display_name_match else pole 
            
        if display_label_b: 
            cmd_text_1 = f"-TEXT {text_x:.3f},{y_pole_main:.3f} 0 {display_label_b}"
            ws.cell(row=row, column=25).value = cmd_text_1
            cmds.append('(command "-layer" "m" "POLE NUMBER" "")')
            cmds.append(cmd_text_1)
        
        # 5. Teks Old Pole / N/A
        if old_pole_text: 
            cmd_text_2 = f"-TEXT {text_x:.3f},{y_pole_lama:.3f} 0 {old_pole_text}"
            ws.cell(row=row, column=26).value = cmd_text_2
            cmds.append('(command "-layer" "m" "TIANG LAMA" "")')
            cmds.append(cmd_text_2)
            
        ws.cell(row=row, column=27).value = data['summary']
            
        # 7. INSERT BLOCK BLACKBOX BK 160A (BBOX 160A)
        for i in range(qty_bk):
            y_insert = bb_start_y - (i * offset_y_bb)
            cmd_bk = f'(command "-insert" "BK" "{x_bk:.3f},{y_insert:.3f}" 1 1 0)'
            cmds.append('(command "-layer" "m" "BBOX 160A" "")')
            cmds.append(cmd_bk)
            
        # 8. INSERT BLOCK BLACKBOX BM 400A (BBOX 400A)
        for i in range(qty_bm):
            y_insert = bb_start_y - (i * offset_y_bb)
            cmd_bm = f'(command "-insert" "BM" "{x_bm:.3f},{y_insert:.3f}" 1 1 0)'
            cmds.append('(command "-layer" "m" "BBOX 400A" "")')
            cmds.append(cmd_bm)

    master_script_lines = []
    
    # Finalisasi dan eksekusi
    for area_name, cmds in scripts_by_area.items():
        cmds.append('(setq ss (ssadd))')
        cmds.append('(setq en firstEnt)')
        cmds.append('(while (setq en (if en (entnext en) (entnext))) (ssadd en ss))')
        cmds.append('(if (> (sslength ss) 0) (command "-group" "c" "*" "Auto-Grouped by Script" ss ""))')
        
        cmds.append('(command "-layer" "s" "0" "")')
        cmds.append('(setvar "OSMODE" 15359)')
        
        file_path = os.path.join(output_folder, f"{area_name}.scr")
        with open(file_path, "w") as scr_file:
            scr_file.write("\n".join(cmds) + "\n")
            
        master_script_lines.extend(cmds)

    master_path = os.path.join(output_folder, "00_RUN_ALL.scr")
    with open(master_path, "w") as master_file:
        master_file.write("\n".join(master_script_lines) + "\n")

    wb.save(output_excel)
    print(f"Beres bro! Routing layer untuk Servis, BK, dan BM udah di-update sesuai request.")

# --- EXECUTION ---
process_and_route("data_input.xlsx", "data_output.xlsx", "output src")